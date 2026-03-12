#!/usr/bin/env python3
"""
BASELINE vs CANDIDATE Ordering Test
======================================
เปรียบเทียบ ordering ของ search results ระหว่าง:
  - BASELINE  : ai-raas-api / text_search (POST + Authorization)
  - CANDIDATE : ai-universal-service-new / placements (GET)

วิธีใช้งาน:
  pip install requests openpyxl
  python run_baseline_vs_candidate.py

หมายเหตุ: ต้องเชื่อมต่อ VPN หรืออยู่ใน network ที่เข้าถึง preprod ได้
"""

import requests
import json
import time
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("กรุณาติดตั้ง openpyxl ก่อน: pip install openpyxl")
    exit(1)

# ===================================================================
# ⚙️  CONFIG — แก้ค่าตรงนี้ถ้าต้องการ
# ===================================================================

# BASELINE endpoint  (ai-raas-api / text_search)
BASELINE_URL = (
    "https://ai-raas-api.trueid-preprod.net"
    "/personalize-rcom/v2/search-api/api/v5/text_search"
)
BASELINE_HEADERS = {
    "Content-Type":  "application/json",
    "Accept":        "application/json, text/plain, */*",
    "Authorization": "5b18e526f0463656f7c4329f90b7ecef9dc546aeb6adad28e911ba82",
}

# CANDIDATE endpoint
CANDIDATE_URL = (
    "http://ai-universal-service-new.preprod-gcp-ai-bn.int-ai-platform.gcp.dmp.true.th"
    "/api/v1/universal/text_search"
)
CANDIDATE_PARAMS_BASE = {
    "userId":   "1",
    "pseudoId": "1",
    "cursor":   "1",
    "limit":    "100",
    "top_k":    "450",
    "ssoId":    "101316473",
}

# Compare top N items
TOP_N = 100

# Rate limiting (seconds between requests)
DELAY = 1

# ===================================================================
# Test matrix — keywords & types
# ===================================================================
KEYWORDS = [
    "มวยไทย",
    " joker",
    " Joker",
    " JOKER",
    " 1",
    " @",
    " ",
    "๒",
    "/",
]

TYPES = [
    "movie",
    "series",
    "livetv",
    "top_results",
    "watch",
    "sfv",
    "privilege",
    "sfvseries",
    "read",
    "channel",
    "ecommerce",
    # "livecommerce",
    "game",
]


def build_test_cases():
    """สร้าง test cases ตามลำดับใน spec"""
    cases = []
    for kw in KEYWORDS:
        for t in ["movie","series","livetv","top_results","watch","sfv","privilege","sfvseries","read","channel","ecommerce","game"]:
            cases.append((kw, t))
    return cases


# ===================================================================
# API CALLERS
# ===================================================================
def call_baseline(keyword: str, type_val: str) -> dict:
    """POST to BASELINE (ai-raas-api) and return ids"""
    body = {
        "debug":           False,
        "no_cache_vector": False,
        "search_keyword":  keyword,
        "top_k":           "450",
        "type":            type_val,
    }
    try:
        resp = requests.post(BASELINE_URL, headers=BASELINE_HEADERS,
                             json=body, timeout=30)
        rj = resp.json()
        # ลอง key ที่เป็นไปได้
        items = None
        for key in ("search_results", "results", "items", "data"):
            if key in rj and isinstance(rj[key], list):
                items = rj[key]
                break
        if items is None and isinstance(rj, list):
            items = rj
        items = items or []
        ids = [x["id"] for x in items if x.get("id")][:TOP_N]
        return {"status": resp.status_code, "ids": ids,
                "count": len(ids), "error": "", "raw": str(rj)[:400]}
    except Exception as e:
        return {"status": "ERROR", "ids": [], "count": 0,
                "error": str(e)[:200], "raw": ""}


def call_candidate(keyword: str, type_val: str) -> dict:
    """GET CANDIDATE and return (status_code, ids, error)"""
    params = {
        **CANDIDATE_PARAMS_BASE,
        "search_keyword": keyword,
        "type":           type_val,
    }
    try:
        resp = requests.get(CANDIDATE_URL, params=params, timeout=30)
        rj = resp.json()
        ids = [x["id"] for x in (rj.get("items") or [])
               if x.get("id")][:TOP_N]
        return {"status": resp.status_code, "ids": ids,
                "count": len(ids), "error": "", "raw": str(rj)[:400]}
    except Exception as e:
        return {"status": "ERROR", "ids": [], "count": 0,
                "error": str(e)[:200], "raw": ""}


# ===================================================================
# SINGLE TEST
# ===================================================================
def run_test(keyword: str, type_val: str, idx: int, total: int) -> dict:
    test_name = (f'CANDIDATE matches BASELINE ordering '
                 f'type : {type_val} (top {TOP_N})')

    print(f"[{idx:03d}/{total}] kw={repr(keyword):<14} type={type_val:<14}", end="")

    # 1. BASELINE
    t0 = time.time()
    bl = call_baseline(keyword, type_val)
    bl_ms = int((time.time() - t0) * 1000)
    time.sleep(DELAY)

    # 2. CANDIDATE
    t0 = time.time()
    cd = call_candidate(keyword, type_val)
    cd_ms = int((time.time() - t0) * 1000)

    # 3. Compare
    if bl["status"] != 200:
        passed = False
        fail_reason = f"BASELINE HTTP {bl['status']}: {bl['error']}"
    elif cd["status"] != 200:
        passed = False
        fail_reason = f"CANDIDATE HTTP {cd['status']}: {cd['error']}"
    elif not bl["ids"] and not cd["ids"]:
        # ทั้งคู่ไม่มีผลลัพธ์ → ถือว่าผ่าน (data not found เหมือนกัน)
        passed = True
        fail_reason = ""
    elif not bl["ids"]:
        passed = False
        fail_reason = "BASELINE returned 0 results (CANDIDATE has results)"
    elif not cd["ids"]:
        passed = False
        fail_reason = "CANDIDATE returned 0 results (BASELINE has results)"
    else:
        # Compare ordering up to TOP_N
        n = min(TOP_N, len(bl["ids"]), len(cd["ids"]))
        passed = (bl["ids"][:n] == cd["ids"][:n])
        if not passed:
            # Find first difference
            mismatches = [(i + 1, bl["ids"][i], cd["ids"][i])
                          for i in range(n) if bl["ids"][i] != cd["ids"][i]]
            first_diff = mismatches[0] if mismatches else ("?", "?", "?")
            fail_reason = (f"{len(mismatches)}/{n} mismatches; "
                           f"first diff at pos {first_diff[0]}: "
                           f"BL={first_diff[1]} vs CD={first_diff[2]}")
        else:
            fail_reason = ""

    label = "✅ PASS" if passed else "❌ FAIL"
    print(f"→ {label}  BL={bl_ms}ms({bl['count']}r) CD={cd_ms}ms({cd['count']}r)")

    # Count matching positions
    if bl["ids"] and cd["ids"]:
        n = min(TOP_N, len(bl["ids"]), len(cd["ids"]))
        matched_positions = sum(1 for i in range(n) if bl["ids"][i] == cd["ids"][i])
    else:
        matched_positions = 0

    return {
        "test_no":           idx,
        "test_name":         test_name,
        "search_keyword":    keyword,
        "type":              type_val,
        # BASELINE
        "bl_status":         bl["status"],
        "bl_ms":             bl_ms,
        "bl_count":          bl["count"],
        "bl_ids_top5":       str(bl["ids"][:5]),
        "bl_error":          bl["error"],
        # CANDIDATE
        "cd_status":         cd["status"],
        "cd_ms":             cd_ms,
        "cd_count":          cd["count"],
        "cd_ids_top5":       str(cd["ids"][:5]),
        "cd_error":          cd["error"],
        # Result
        "passed":            passed,
        "fail_reason":       fail_reason,
        "matched_positions": matched_positions,
        "compared_n":        min(TOP_N, len(bl["ids"]), len(cd["ids"])),
    }


# ===================================================================
# EXCEL REPORT
# ===================================================================
def create_excel_report(results: list, output_path: str):
    wb = openpyxl.Workbook()

    total   = len(results)
    passed  = sum(1 for r in results if r["passed"])
    failed  = total - passed

    # ── Palette ─────────────────────────────────────────────────────
    C_BLUE   = "1565C0"
    C_WHITE  = "FFFFFF"
    C_GREEN  = "00C851"
    C_RED    = "FF4444"
    C_ORANGE = "FF6D00"
    C_LGRE   = "E8F5E9"
    C_LYEL   = "FFFDE7"
    C_LBLUE  = "E3F2FD"

    def fill(h): return PatternFill(start_color=h, end_color=h, fill_type="solid")
    def bd():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)
    center = Alignment(horizontal="center", vertical="center")
    wrap   = Alignment(vertical="center", wrap_text=True)

    # ================================================================
    # Sheet 1 – Test Results
    # ================================================================
    ws = wb.active
    ws.title = "Test Results"

    # Title
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value     = "🔍  BASELINE vs CANDIDATE Ordering Test Report"
    c.font      = Font(size=15, bold=True, color=C_WHITE)
    c.fill      = fill(C_BLUE)
    c.alignment = center
    ws.row_dimensions[1].height = 36

    # Meta
    meta = [
        ("Run Date",     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
         "Top N Compared", str(TOP_N)),
        ("Total Tests",  total,   "Passed",    passed),
        ("Failed",       failed,  "Pass Rate",
         f"{passed/total*100:.1f}%" if total else "0%"),
    ]
    for ri, (k1, v1, k2, v2) in enumerate(meta, 2):
        for col, val in [(1, k1), (2, v1), (5, k2), (6, v2)]:
            cell = ws.cell(row=ri, column=col, value=val)
            if col in (1, 5): cell.font = Font(bold=True)
    ws.cell(3, 2).font = Font(bold=True, color=C_GREEN if passed == total else C_ORANGE)
    ws.cell(4, 2).font = Font(bold=True, color=C_RED if failed > 0 else C_GREEN)

    # Stat boxes
    for col, (lbl, val, color) in enumerate([
        ("TOTAL",  total,  C_BLUE),
        ("PASSED", passed, C_GREEN),
        ("FAILED", failed, C_RED if failed > 0 else "9E9E9E"),
    ], start=10):
        lc = ws.cell(2, col, lbl)
        vc = ws.cell(3, col, val)
        for cell in (lc, vc):
            cell.fill = fill(color)
            cell.font = Font(bold=True, color=C_WHITE,
                             size=12 if cell == vc else 9)
            cell.alignment = center
            cell.border = bd()
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Headers
    HDR = 5
    headers = [
        "#", "Test Name (keyword / type)",
        "Keyword", "Type",
        "BL Status", "BL Time(ms)", "BL Results",
        "CD Status", "CD Time(ms)", "CD Results",
        "Matched / N", "Result", "Fail Reason",
        "BASELINE top-5 IDs",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(HDR, col, h)
        cell.fill = fill(C_BLUE)
        cell.font = Font(bold=True, color=C_WHITE)
        cell.alignment = center
        cell.border = bd()
    ws.row_dimensions[HDR].height = 24

    # Data
    for r in results:
        row = HDR + r["test_no"]
        rf  = fill(C_LGRE) if r["passed"] else fill(C_LYEL)
        matched_label = (f"{r['matched_positions']}/{r['compared_n']}"
                         if r["compared_n"] else "–")
        vals = [
            r["test_no"],
            f'{repr(r["search_keyword"])} / {r["type"]}',
            r["search_keyword"],
            r["type"],
            r["bl_status"], r["bl_ms"], r["bl_count"],
            r["cd_status"], r["cd_ms"], r["cd_count"],
            matched_label,
            "PASS" if r["passed"] else "FAIL",
            r["fail_reason"],
            r["bl_ids_top5"],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row, col, val)
            cell.fill      = rf
            cell.border    = bd()
            cell.alignment = wrap if col in (2, 13, 14) else center
            if col == 12:
                cell.font = Font(bold=True,
                                 color=C_GREEN if r["passed"] else C_RED)
            if col == 11 and r["compared_n"]:
                pct = r["matched_positions"] / r["compared_n"]
                cell.font = Font(bold=True,
                                 color=C_GREEN  if pct == 1.0
                                 else  C_ORANGE if pct >= 0.8
                                 else  C_RED)
        ws.row_dimensions[row].height = 18

    # Column widths
    for i, w in enumerate([
        4, 32, 14, 14, 10, 11, 10, 10, 11, 10, 12, 8, 50, 50
    ], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{HDR+1}"

    # ================================================================
    # Sheet 2 – By Type
    # ================================================================
    ws2 = wb.create_sheet("By Type")
    ws2.merge_cells("A1:F1")
    c = ws2["A1"]
    c.value = "Results by Type"
    c.font  = Font(size=13, bold=True, color=C_WHITE)
    c.fill  = fill(C_BLUE)
    c.alignment = center

    for col, h in enumerate(["Type","Total","Passed","Failed","Pass Rate","Avg Match%"], 1):
        cell = ws2.cell(2, col, h)
        cell.fill = fill(C_LBLUE); cell.font = Font(bold=True)
        cell.alignment = center; cell.border = bd()

    type_stats = {}
    for r in results:
        s = type_stats.setdefault(r["type"], {"total":0,"passed":0,"match":[]})
        s["total"] += 1
        if r["passed"]: s["passed"] += 1
        if r["compared_n"]:
            s["match"].append(r["matched_positions"] / r["compared_n"])

    for ri, (t, s) in enumerate(type_stats.items(), 3):
        fc   = s["total"] - s["passed"]
        rate = f"{s['passed']/s['total']*100:.0f}%" if s["total"] else "–"
        avg_m = f"{sum(s['match'])/len(s['match'])*100:.1f}%" if s["match"] else "–"
        for col, val in enumerate([t, s["total"], s["passed"], fc, rate, avg_m], 1):
            cell = ws2.cell(ri, col, val)
            cell.border = bd(); cell.alignment = center
            if col == 3: cell.font = Font(color=C_GREEN, bold=True)
            if col == 4 and fc > 0: cell.font = Font(color=C_RED, bold=True)
            if col == 5:
                cell.font = Font(bold=True,
                    color=C_GREEN if fc==0 else (C_RED if fc==s["total"] else C_ORANGE))

    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    # ================================================================
    # Sheet 3 – By Keyword
    # ================================================================
    ws3 = wb.create_sheet("By Keyword")
    ws3.merge_cells("A1:F1")
    c = ws3["A1"]
    c.value = "Results by Keyword"
    c.font  = Font(size=13, bold=True, color=C_WHITE)
    c.fill  = fill(C_BLUE)
    c.alignment = center

    for col, h in enumerate(["Keyword (repr)","Total","Passed","Failed","Pass Rate","Avg Match%"], 1):
        cell = ws3.cell(2, col, h)
        cell.fill = fill(C_LBLUE); cell.font = Font(bold=True)
        cell.alignment = center; cell.border = bd()

    kw_stats = {}
    for r in results:
        s = kw_stats.setdefault(r["search_keyword"], {"total":0,"passed":0,"match":[]})
        s["total"] += 1
        if r["passed"]: s["passed"] += 1
        if r["compared_n"]:
            s["match"].append(r["matched_positions"] / r["compared_n"])

    for ri, (kw, s) in enumerate(kw_stats.items(), 3):
        fc   = s["total"] - s["passed"]
        rate = f"{s['passed']/s['total']*100:.0f}%" if s["total"] else "–"
        avg_m = f"{sum(s['match'])/len(s['match'])*100:.1f}%" if s["match"] else "–"
        for col, val in enumerate([repr(kw), s["total"], s["passed"], fc, rate, avg_m], 1):
            cell = ws3.cell(ri, col, val)
            cell.border = bd(); cell.alignment = center
            if col == 3: cell.font = Font(color=C_GREEN, bold=True)
            if col == 4 and fc > 0: cell.font = Font(color=C_RED, bold=True)
            if col == 5:
                cell.font = Font(bold=True,
                    color=C_GREEN if fc==0 else (C_RED if fc==s["total"] else C_ORANGE))

    ws3.column_dimensions["A"].width = 20
    for col in range(2, 7):
        ws3.column_dimensions[get_column_letter(col)].width = 15

    # ================================================================
    # Sheet 4 – Mismatch Detail
    # ================================================================
    ws4 = wb.create_sheet("Mismatch Detail")
    ws4.merge_cells("A1:D1")
    c = ws4["A1"]
    c.value = "Failed Test Details"
    c.font  = Font(size=13, bold=True, color=C_WHITE)
    c.fill  = fill(C_RED)
    c.alignment = center

    for col, h in enumerate(["#","keyword / type","Fail Reason","BASELINE top-5 IDs"], 1):
        cell = ws4.cell(2, col, h)
        cell.fill = fill(C_LBLUE); cell.font = Font(bold=True)
        cell.border = bd()

    failed_rows = [r for r in results if not r["passed"]]
    if failed_rows:
        for ri, r in enumerate(failed_rows, 3):
            vals = [
                r["test_no"],
                f'{repr(r["search_keyword"])} / {r["type"]}',
                r["fail_reason"],
                r["bl_ids_top5"],
            ]
            for col, val in enumerate(vals, 1):
                cell = ws4.cell(ri, col, val)
                cell.fill = fill(C_LYEL)
                cell.border = bd()
                cell.alignment = wrap if col in (2,3,4) else center
            ws4.row_dimensions[ri].height = 40
    else:
        ws4.cell(3, 1, "🎉 All tests passed!").font = Font(color=C_GREEN, bold=True, size=12)

    ws4.column_dimensions["A"].width = 5
    ws4.column_dimensions["B"].width = 32
    ws4.column_dimensions["C"].width = 60
    ws4.column_dimensions["D"].width = 55

    # ── Save ─────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"\n✅ Excel report saved → {output_path}")


# ===================================================================
# MAIN
# ===================================================================
def main():
    print("=" * 72)
    print("  BASELINE vs CANDIDATE Ordering Test Runner")
    print(f"  BASELINE  : {BASELINE_URL}")
    print(f"  CANDIDATE : {CANDIDATE_URL}")
    print(f"  Top N     : {TOP_N}")
    print(f"  Start     : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 72)

    cases = build_test_cases()
    total = len(cases)
    print(f"  Total test cases: {total}  (each case = 2 API calls)\n")

    results = []
    for i, (kw, t) in enumerate(cases, 1):
        result = run_test(kw, t, i, total)
        results.append(result)
        time.sleep(DELAY)

    passed = sum(1 for r in results if r["passed"])
    failed = total - passed

    print("\n" + "=" * 72)
    print(f"  FINAL : {passed}/{total} PASSED  |  {failed} FAILED")
    print(f"  Rate  : {passed/total*100:.1f}%")
    print("=" * 72)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path  = os.path.join(script_dir, "Baseline_vs_Candidate_Report.xlsx")
    json_path  = os.path.join(script_dir, "Baseline_vs_Candidate_Results.json")

    create_excel_report(results, xlsx_path)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON results saved → {json_path}")


if __name__ == "__main__":
    main()