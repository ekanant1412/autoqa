#!/usr/bin/env python3
"""
Atlas Serving Endpoint Comparison Script
Compare response bodies between 2 endpoints for all placements.

Usage:
    pip install requests openpyxl pandas deepdiff
    python compare_endpoints.py

Output:
    - comparison_results.xlsx  : Full result with diff details
    - comparison_summary.txt   : Quick summary
"""

import requests
import pandas as pd
import json
import time
import sys
from deepdiff import DeepDiff
from datetime import datetime

# ─── CONFIG ─────────────────────────────────────────────────────────────────
ENDPOINT_1 = "http://atlas-serving.preprod-gcp-ai-bn.int-ai-platform.gcp.dmp.true.th"
ENDPOINT_2 = "http://atlas-serving-2.prod-gcp-ai-bn.ai-platform.gcp.dmp.true.th"

EXCEL_INPUT = "/Users/ekananat/Desktop/autoqa/src/req_atlas.xlsx" 
EXCEL_OUTPUT = "comparison_results.xlsx"
SUMMARY_OUTPUT = "comparison_summary.txt"

TIMEOUT = 10       # seconds per request
DELAY   = 0.15     # seconds between requests (avoid rate limiting)
# ─────────────────────────────────────────────────────────────────────────────


def fetch(url: str) -> dict:
    """Call URL, return dict with status, items_count, body (parsed JSON or None), error."""
    try:
        resp = requests.get(url, timeout=TIMEOUT)
        status = resp.status_code
        body = None
        items_count = 0
        try:
            body = resp.json()
            items = body.get("items") or body.get("data", {}).get("items", [])
            items_count = len(items) if isinstance(items, list) else 0
        except Exception:
            body = None
        return {"status": status, "items_count": items_count, "body": body, "error": None}
    except requests.exceptions.ConnectionError as e:
        return {"status": "ERROR", "items_count": 0, "body": None, "error": f"ConnectionError: {e}"}
    except requests.exceptions.Timeout:
        return {"status": "TIMEOUT", "items_count": 0, "body": None, "error": "Request timed out"}
    except Exception as e:
        return {"status": "ERROR", "items_count": 0, "body": None, "error": str(e)}


def compare_bodies(body1, body2) -> dict:
    """
    Deep-compare two JSON bodies.
    Returns:
        match     : True/False/None (None = both failed)
        diff_type : 'SAME' | 'BODY_DIFF' | 'BOTH_FAILED' | 'ONE_FAILED' | 'ITEMS_COUNT_DIFF'
        diff_detail: human-readable diff summary
    """
    if body1 is None and body2 is None:
        return {"match": None, "diff_type": "BOTH_FAILED", "diff_detail": "Both endpoints returned no parseable body"}

    if body1 is None:
        return {"match": False, "diff_type": "ONE_FAILED", "diff_detail": "Endpoint 1 returned no parseable body"}

    if body2 is None:
        return {"match": False, "diff_type": "ONE_FAILED", "diff_detail": "Endpoint 2 returned no parseable body"}

    # Compare items array specifically (most important)
    items1 = body1.get("items") or body1.get("data", {}).get("items", [])
    items2 = body2.get("items") or body2.get("data", {}).get("items", [])
    items1 = items1 if isinstance(items1, list) else []
    items2 = items2 if isinstance(items2, list) else []

    # Deep compare full body
    try:
        diff = DeepDiff(body1, body2, ignore_order=True, significant_digits=2,
                exclude_paths=["root['schemaId']", "root['request_id']"])
        if not diff:
            return {"match": True, "diff_type": "SAME", "diff_detail": ""}

        # Summarize diff
        diff_parts = []
        if "values_changed" in diff:
            changed_details = []
            for key, change in list(diff["values_changed"].items())[:3]:
                old_val = change.get("old_value", "?")
                new_val = change.get("new_value", "?")
                changed_details.append(f"{key}({old_val!r} → {new_val!r})")
            diff_parts.append(f"values_changed: {changed_details}")
        if "iterable_item_added" in diff:
            diff_parts.append(f"items_added in ep2: {len(diff['iterable_item_added'])}")
        if "iterable_item_removed" in diff:
            diff_parts.append(f"items_removed in ep2: {len(diff['iterable_item_removed'])}")
        if "dictionary_item_added" in diff:
            diff_parts.append(f"keys_added: {list(diff['dictionary_item_added'])[:3]}")
        if "dictionary_item_removed" in diff:
            diff_parts.append(f"keys_removed: {list(diff['dictionary_item_removed'])[:3]}")
        if "type_changes" in diff:
            diff_parts.append(f"type_changes: {list(diff['type_changes'].keys())[:3]}")

        # Check if only item count differs
        if len(items1) != len(items2):
            diff_type = "ITEMS_COUNT_DIFF"
        else:
            diff_type = "BODY_DIFF"

        return {
            "match": False,
            "diff_type": diff_type,
            "diff_detail": " | ".join(diff_parts) if diff_parts else str(diff)[:200]
        }
    except Exception as e:
        # Fallback: simple equality
        match = (body1 == body2)
        return {
            "match": match,
            "diff_type": "SAME" if match else "BODY_DIFF",
            "diff_detail": "" if match else f"diff-error: {e}"
        }


def main():
    print(f"\n{'='*60}")
    print(f"Atlas Endpoint Comparison")
    print(f"Endpoint 1: {ENDPOINT_1}")
    print(f"Endpoint 2: {ENDPOINT_2}")
    print(f"{'='*60}\n")

    # Load placements from Excel
    try:
        df = pd.read_excel(EXCEL_INPUT)
        print(f"✅ Loaded {len(df)} placements from {EXCEL_INPUT}")
    except FileNotFoundError:
        print(f"❌ File not found: {EXCEL_INPUT}")
        sys.exit(1)

    HOST1 = ENDPOINT_1.replace("http://", "")
    HOST2 = ENDPOINT_2.replace("http://", "")

    results = []
    total = len(df)

    for i, row in df.iterrows():
        no        = row["No"]
        placement = str(row["Placement"]).strip()
        url1      = str(row["URL"]).strip()
        url2      = url1.replace(HOST1, HOST2)

        sys.stdout.write(f"\r[{no:3d}/{total}] {placement[:55]:<55}")
        sys.stdout.flush()

        # Fetch both endpoints
        r1 = fetch(url1)
        time.sleep(DELAY)
        r2 = fetch(url2)
        time.sleep(DELAY)

        # Compare
        cmp = compare_bodies(r1["body"], r2["body"])

        # Status match (HTTP status code)
        status_match = (r1["status"] == r2["status"])

        results.append({
            "No":             no,
            "Placement":      placement,
            # Endpoint 1
            "EP1_Status":     r1["status"],
            "EP1_Items":      r1["items_count"],
            "EP1_Error":      r1["error"] or "",
            # Endpoint 2
            "EP2_Status":     r2["status"],
            "EP2_Items":      r2["items_count"],
            "EP2_Error":      r2["error"] or "",
            # Comparison
            "Status_Match":   "✅" if status_match else "❌",
            "Body_Match":     "✅" if cmp["match"] is True else ("⬛" if cmp["match"] is None else "❌"),
            "Diff_Type":      cmp["diff_type"],
            "Diff_Detail":    cmp["diff_detail"],
            "URL1":           url1,
            "URL2":           url2,
        })

    print(f"\n\n✅ Done! Collected {len(results)} results.\n")

    # ── Build result DataFrame ──────────────────────────────────────────────
    result_df = pd.DataFrame(results)

    # ── Summary stats ───────────────────────────────────────────────────────
    same       = result_df[result_df["Diff_Type"] == "SAME"].shape[0]
    body_diff  = result_df[result_df["Diff_Type"] == "BODY_DIFF"].shape[0]
    cnt_diff   = result_df[result_df["Diff_Type"] == "ITEMS_COUNT_DIFF"].shape[0]
    one_fail   = result_df[result_df["Diff_Type"] == "ONE_FAILED"].shape[0]
    both_fail  = result_df[result_df["Diff_Type"] == "BOTH_FAILED"].shape[0]
    status_mismatch = result_df[result_df["Status_Match"] == "❌"].shape[0]

    summary = f"""
Atlas Endpoint Comparison Summary
Run at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Endpoint 1: {ENDPOINT_1}
Endpoint 2: {ENDPOINT_2}
Total placements: {total}

── Body Comparison ──────────────────
  SAME              : {same}
  BODY_DIFF         : {body_diff}
  ITEMS_COUNT_DIFF  : {cnt_diff}
  ONE_FAILED        : {one_fail}
  BOTH_FAILED       : {both_fail}

── Status Code ──────────────────────
  Status mismatch   : {status_mismatch}

── Placements with differences ──────
"""
    diff_rows = result_df[result_df["Diff_Type"] != "SAME"]
    for _, r in diff_rows.iterrows():
        summary += (
            f"  [{r['No']:3d}] {r['Placement']:<50} "
            f"EP1={r['EP1_Status']}({r['EP1_Items']}) "
            f"EP2={r['EP2_Status']}({r['EP2_Items']}) "
            f"→ {r['Diff_Type']}\n"
        )

    print(summary)

    # Save summary
    with open(SUMMARY_OUTPUT, "w") as f:
        f.write(summary)
    print(f"📄 Summary saved: {SUMMARY_OUTPUT}")

    # ── Export Excel ────────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_all   = wb.active
    ws_all.title = "All Results"
    ws_diff  = wb.create_sheet("Differences")
    ws_same  = wb.create_sheet("Same")

    COLS = [
        "No", "Placement",
        "EP1_Status", "EP1_Items",
        "EP2_Status", "EP2_Items",
        "Status_Match", "Body_Match", "Diff_Type", "Diff_Detail",
        "URL1", "URL2"
    ]

    # Color helpers
    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    RED    = PatternFill("solid", fgColor="FFC7CE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    GREY   = PatternFill("solid", fgColor="D9D9D9")
    HEADER = PatternFill("solid", fgColor="366092")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
    BOLD   = Font(bold=True, name="Arial")
    NORMAL = Font(name="Arial", size=10)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_header(ws, cols):
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30

    def write_row(ws, row_num, row_data, cols, fill=None):
        diff_type = row_data.get("Diff_Type", "")
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=row_num, column=ci, value=row_data.get(col, ""))
            cell.font = NORMAL
            cell.alignment = Alignment(wrap_text=(col in ("Diff_Detail",)), vertical="center")
            cell.border = border
            # Row color
            if fill:
                cell.fill = fill
            elif diff_type == "SAME":
                cell.fill = GREEN
            elif diff_type in ("BODY_DIFF", "ONE_FAILED"):
                cell.fill = RED
            elif diff_type == "ITEMS_COUNT_DIFF":
                cell.fill = YELLOW
            elif diff_type == "BOTH_FAILED":
                cell.fill = GREY

    def set_col_widths(ws, widths):
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

    # ── Write All Results sheet ─────────────────────────────────────────────
    write_header(ws_all, COLS)
    for ri, row_data in enumerate(results, 2):
        write_row(ws_all, ri, row_data, COLS)

    set_col_widths(ws_all, {
        "A": 5, "B": 45, "C": 12, "D": 10,
        "E": 12, "F": 10, "G": 14, "H": 12,
        "I": 20, "J": 50, "K": 80, "L": 80
    })
    ws_all.freeze_panes = "A2"

    # ── Write Differences sheet ─────────────────────────────────────────────
    write_header(ws_diff, COLS)
    diff_data = [r for r in results if r["Diff_Type"] != "SAME"]
    for ri, row_data in enumerate(diff_data, 2):
        write_row(ws_diff, ri, row_data, COLS)
    set_col_widths(ws_diff, {
        "A": 5, "B": 45, "C": 12, "D": 10,
        "E": 12, "F": 10, "G": 14, "H": 12,
        "I": 20, "J": 50, "K": 80, "L": 80
    })
    ws_diff.freeze_panes = "A2"

    # ── Write Same sheet ────────────────────────────────────────────────────
    write_header(ws_same, COLS)
    same_data = [r for r in results if r["Diff_Type"] == "SAME"]
    for ri, row_data in enumerate(same_data, 2):
        write_row(ws_same, ri, row_data, COLS)
    set_col_widths(ws_same, {
        "A": 5, "B": 45, "C": 12, "D": 10,
        "E": 12, "F": 10, "G": 14, "H": 12,
        "I": 20, "J": 50, "K": 80, "L": 80
    })
    ws_same.freeze_panes = "A2"

    # ── Summary sheet ───────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary", 0)
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 15

    def sum_row(r, label, value, fill=None):
        c1 = ws_sum.cell(row=r, column=1, value=label)
        c2 = ws_sum.cell(row=r, column=2, value=value)
        c1.font = BOLD
        c2.font = NORMAL
        c2.alignment = Alignment(horizontal="center")
        if fill:
            c1.fill = fill
            c2.fill = fill

    ws_sum.cell(row=1, column=1, value="Comparison Summary").font = Font(bold=True, size=14, name="Arial")
    ws_sum.cell(row=2, column=1, value=f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = NORMAL
    ws_sum.cell(row=3, column=1, value=f"EP1: {ENDPOINT_1}").font = NORMAL
    ws_sum.cell(row=4, column=1, value=f"EP2: {ENDPOINT_2}").font = NORMAL
    sum_row(6,  "Total Placements",    total)
    sum_row(7,  "✅ SAME",             same,       GREEN)
    sum_row(8,  "❌ BODY_DIFF",        body_diff,  RED)
    sum_row(9,  "⚠️  ITEMS_COUNT_DIFF", cnt_diff,   YELLOW)
    sum_row(10, "⚠️  ONE_FAILED",       one_fail,   YELLOW)
    sum_row(11, "⬛ BOTH_FAILED",      both_fail,  GREY)
    sum_row(13, "❌ Status Mismatch",  status_mismatch, RED)

    wb.save(EXCEL_OUTPUT)
    print(f"📊 Excel saved: {EXCEL_OUTPUT}")
    print(f"\n🎉 All done!")


if __name__ == "__main__":
    main()
